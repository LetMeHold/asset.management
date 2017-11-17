# -*- coding: utf-8 -*-

from gl import *
from wrap.base import *
from openpyxl import Workbook
from openpyxl import load_workbook

class Business:

    def __init__(self):
        self.db = DB(db='asset.management')
        self.tol = Tools()

    def __del__(self):
        if self.db != None:
            del self.db
            self.db = None

    def getRedPrice(self, vc, typ, spec):
        sql = self.tol.queryVcZhSql(vc)
        vcList = self.db.query(sql)
        if vcList==False or len(vcList)==0:
            GL.setErr('查询电压等级对照表时失败！')
            return False
        GL.LOG.debug('查询出电压对照: %s' % str(vcList))
        if len(vcList) == 1:
            tmp = vcList[0]['vczh']
        else:
            tmp = []
            for m in vcList:
                tmp.append(m['vczh'])
            tmp = tuple(tmp)
        sql = self.tol.queryRedPriceSql(tmp, typ, spec)
        rpList = self.db.query(sql)
        if rpList==False or len(rpList)==0:
            GL.setErr('查询红本单价时失败！')
            return False
        GL.LOG.debug('查询出红本单价: %s' % str(rpList))
        ret = False
        if len(rpList) == 1:
            ret = rpList[0]['price']
        else:
            name = '%s-%s%s %s' % (typ,vc[:vc.find('/')],vcList[0]['unit'],spec)
            for m in rpList:
                if m['name'].lower() == name.lower():
                    ret = m['price']
                    GL.LOG.info('红本价查到多条记录,进一步筛选(%s)成功' % name)
                    break
        if ret == False:
            GL.setErr('红本价查到多条记录,进一步筛选(%s)失败！' % name)
        return ret

    def getClassDiscount(self, classify, sn):
        sql = self.tol.queryClassDiscountSql(classify, sn)
        result = self.db.query(sql)
        if result == False:
            return False
        retLst = []
        for m in result:
            retLst.append(m['discount'])
        if len(retLst) == 0:
            GL.setErr('未获取分类(%s%d)到下浮数据！' % (classify,sn))
            return False
        return retLst

    def getValueRedPrice(self, rp, discount, amount):
        ret = rp
        for n in discount:
            ret *= n
        ret *= amount
        return ret

    def getStuff(self, vc, sn, typ, spec, amount):
        sql = self.tol.queryStuffSql(vc, sn, typ, spec)
        result = self.db.query(sql)
        if result == False:
            return False
        if len(result) != 1:
            GL.setErr('表stuff中查出0条或超过1条记录:%s' % result)
            return False
        result = result[0]
        stuffLst = []
        for i in range(GL.minStuffid,GL.maxStuffid+1):
            index = 'sid%d' % i
            stuffLst.append(result[index]*amount)
        return stuffLst

    def record(self, orderno, orderdate, ordertype, orderspec, vc, typ, spec, classify, sn, amount, outPrice, stuffLst):
        tmp = '(orderno,orderdate,ordertype,orderspec,type,spec,vc,class,sn,amount,outprice'
        for k in range(GL.minStuffid,GL.maxStuffid+1):
            tmp = '%s,sid%d' % (tmp,k)
        tmp += ')'
        lst = [orderno,orderdate,ordertype,orderspec,typ,spec,vc,classify,sn,amount,outPrice/1000]
        for sl in stuffLst:
            lst.append(sl/1000)
        sql = 'insert into record %s values %s' % (tmp,str(tuple(lst)))
        self.db.exec(sql)

    def getStuffNameById(self, index):
        result = self.db.query('select name from stuffid where id = %d' % index)
        if result == False:
            return False
        return result[0]['name']

    def loadClassifyExcel(self):
        wb = load_workbook(filename='../../db/zjh/分类下浮标准对照表.xlsx')
        for ws in wb:
            self.db.resetCount()
            for row in ws.rows:
                if len(row) == 3:
                    clas = row[0].value
                    sn = row[1].value
                    discount = row[2].value
                    values = (clas,sn,discount)
                    sql = 'insert into classify (class,sn,discount) values %s' % str(values)
                    self.db.exec(sql)
            GL.LOG.info('导入分类下浮标准对照表(%s)成功与失败次数: %s' % (ws.title,str(self.db.getCount())))

    def loadVcMapExcel(self):
        wb = load_workbook(filename='../../db/zjh/电压等级对照表.xlsx')
        for ws in wb:
            self.db.resetCount()
            for row in ws.rows:
                if len(row) == 3:
                    vc = str(row[0].value).strip()
                    vczh = row[1].value.strip()
                    unit = row[2].value.strip()
                    values = (vc,vczh,unit)
                    sql = 'insert into vcmap (vc,vczh,unit) values %s' % str(values)
                    self.db.exec(sql)
            GL.LOG.info('导入电压等级对照表(%s)成功与失败次数: %s' % (ws.title,str(self.db.getCount())))

    def loadRedPriceExcel(self):
        wb = load_workbook(filename='../../db/zjh/电子红本2017.3.14.xlsx')
        for ws in wb:
            self.db.resetCount()
            for row in ws.rows:
                if isinstance(row[0].value,int) == False:   #跳过标题行
                    continue
                if len(row) == 6:
                    vc = row[1].value.strip()
                    typ = row[2].value.strip()
                    name = row[3].value.strip()
                    spec = name[name.rfind(' '):].strip()
                    #spec = spec[len(typ):].strip()
                    unit = row[4].value.strip()
                    price = row[5].value
                    values = (vc,typ,name,spec,unit,price)
                    sql = 'insert into redprice (vc,type,name,spec,unit,price) values %s' % str(values)
                    self.db.exec(sql)
            GL.LOG.info('导入红本价(%s)成功与失败次数: %s' % (ws.title,str(self.db.getCount())))

    def loadStuffExcel(self):
        wb = load_workbook(filename='../../db/zjh/用料汇总.xlsx')
        for ws in wb:
            self.db.resetCount()
            for row in ws.rows:
                tl = []
                if len(row)>=28 and isinstance(row[0].value,int):
                    sn = row[0].value
                    clas = row[1].value
                    spec = row[2].value
                    vc = row[3].value
                    tl.extend([sn,clas,spec,vc])
                    for i in range(4,28):
                        m = row[i].value
                        tl.append(m)
                    tmp = '(sn,type,spec,vc'
                    for k in range(1,25):
                        tmp = '%s,sid%d' % (tmp,k)
                    tmp += ')'
                    sql = 'insert into stuff %s values %s' % (tmp,str(tuple(tl)))
                    self.db.exec(sql)
            GL.LOG.info('导入分类下浮标准对照表(%s)成功与失败次数: %s' % (ws.title,str(self.db.getCount())))


